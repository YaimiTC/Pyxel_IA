# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64, io, re, hashlib
from datetime import date, datetime
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except Exception:
    openpyxl = None


def _col_to_idx(col_ref):
    """Convierte 'A'/'B'/'AA' o '1'/'2' a índice 0-based."""
    if isinstance(col_ref, int):
        return max(col_ref - 1, 0)
    s = (col_ref or '').strip()
    if not s:
        raise UserError(_("Las referencias de columna no pueden estar vacías."))
    if s.isdigit():
        return max(int(s) - 1, 0)
    s = s.upper()
    if not re.fullmatch(r'[A-Z]+', s):
        raise UserError(_("Columna inválida: %s (usa letras A,B,AA o número 1,2,3)") % s)
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


# Busca textos tipo: "Periodo: 25/06/2025 - 20/07/2025"
PERIOD_REGEX = re.compile(
    r'Periodo:\s*(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})',
    re.IGNORECASE
)


def _parse_dt_ddmmyyyy(s):
    return datetime.strptime(s, '%d/%m/%Y').date()


class ImportContainerBillingWizard(models.TransientModel):
    _name = 'import.container.billing.wizard'
    _description = 'Cargar Excel para facturar contenedores por importación'

    # ===== Archivo =====
    file_data = fields.Binary(string="Archivo Excel (.xlsx)", required=True)
    file_name = fields.Char(string="Nombre de archivo")

    # ===== Navegación de hoja/encabezado =====
    sheet_index = fields.Integer(string="Hoja (0=primera)", default=0)
    header_row = fields.Integer(
        string="Fila encabezado", default=1,
        help="Número de fila del encabezado (la siguiente es la primera de datos)"
    )

    # ===== Mapeo de columnas =====
    col_container = fields.Char(string="Columna Contenedor", required=True, help="Ej: A o 1")
    col_bl = fields.Char(string="Columna BL", required=False, help="Opcional si el contenedor es único")
    col_cost = fields.Char(string="Columna Costo", required=True)
    col_customer = fields.Char(string="Columna Cliente (opcional)")  # solo para la vista (no se usa)

    # ===== Parámetros de creación =====
    product_id = fields.Many2one(
        'product.product', string="Producto/Servicio a facturar", required=True,
        domain=[('detailed_type', '=', 'service')]
    )
    journal_id = fields.Many2one(
        'account.journal', string="Diario de ventas", required=True,
        domain=[('type', '=', 'sale')]
    )
    company_id = fields.Many2one(
        'res.company', string="Compañía",
        default=lambda self: self.env.company, required=True
    )
    invoice_date = fields.Date(string="Fecha de factura", default=fields.Date.context_today)
    force_currency_id = fields.Many2one('res.currency', string="Forzar moneda",
                                        default=lambda self: self.env.company.currency_id,)

    # Publicación (la vista usa draft_only)
    draft_only = fields.Boolean(string="Crear en borrador", default=True)

    # División/Redondeo (placeholders para la vista)
    price_division_mode = fields.Selection([
        ('equal', 'Partes iguales'),
        ('by_lines', 'Proporcional a líneas'),
    ], string="Modo de división", default='equal')
    rounding_mode = fields.Selection([
        ('half_up', 'Por exceso'),
        ('half_even', 'Por defecto'),
    ], string="Redondeo", default='half_up')

    # Comportamiento de omisión
    skip_missing_container = fields.Boolean(string="Omitir sin contenedor", default=True)
    skip_missing_import = fields.Boolean(string="Omitir sin importación", default=True)

    # Mensaje de resultado
    result_message = fields.Text(string="Resumen")

    # Campo informativo del período detectado en el Excel
    period_start = fields.Date(string="Período desde", readonly=True)
    period_end = fields.Date(string="Período hasta", readonly=True)

    grouping_mode = fields.Selection(
        [
            ('by_import', 'Agrupar por Cliente + Importación (actual)'),
            ('by_partner', 'Agrupar solo por Cliente'),
        ],
        string="Modo de agrupación",
        default='by_import',
        help="• Por defecto se agrupa por (Cliente + Importación) y se vinculan contenedores.\n"
             "• Si eliges 'Agrupar solo por Cliente', todas las líneas de ese cliente irán a una sola factura "
             "sin setear x_studio_import_id ni x_studio_container_ids."
    )

    # Alias del botón de la vista
    def action_process_file(self):
        return self.action_process()

    # ===== Aliases de modelos/campos personalizados =====
    _container_model = 'x_container'
    _import_model = 'x_import'
    _container_name_field = 'x_name'
    _container_bl_field = 'x_studio_bill_of_landing_number'  # BL en la importación
    _container_import_m2o = 'x_studio_import_id'  # M2O hacia importación

    # O2M intermedio y su M2O a PO line
    _container_lines_o2m = 'x_studio_order_lines_by_container'  # -> x_containerlineorder (intermedio)
    _container_line_po_line_m2o = 'x_studio_purchase_order_line'  # -> purchase.order.line

    # Campos custom en factura
    _move_import_m2o = 'x_studio_import_id'
    _move_container_m2m = 'x_studio_container_ids'  # M2M
    _move_type_import_field = 'x_studio_type_import'  # 'normal'

    # ====================== Core ========================
    def _get_worksheet(self, wb):
        """Obtiene la worksheet por índice con validación."""
        idx = self.sheet_index or 0
        sheets = wb.worksheets
        if idx < 0 or idx >= len(sheets):
            raise UserError(_("Índice de hoja fuera de rango. El archivo tiene %s hoja(s).") % len(sheets))
        return sheets[idx]

    def _parse_excel(self):
        """Devuelve [{'row': n, 'container':..., 'bl':..., 'cost': float}, ...]"""
        if openpyxl is None:
            raise UserError(_("Falta la librería 'openpyxl'. Instálala en el servidor."))

        if not self.file_data:
            raise UserError(_("Adjunta un archivo Excel."))

        try:
            data = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("No se pudo leer el Excel: %s") % e)

        ws = self._get_worksheet(wb)

        c_container = _col_to_idx(self.col_container)
        c_bl = _col_to_idx(self.col_bl) if self.col_bl else None
        c_cost = _col_to_idx(self.col_cost)

        start_row = max(self.header_row or 1, 1) + 1  # datos empiezan debajo del encabezado
        results = []

        for r in range(start_row, ws.max_row + 1):
            def _val(ci):
                if ci is None:
                    return ''
                v = ws.cell(row=r, column=ci + 1).value
                return '' if v is None else str(v).strip()

            raw_container = _val(c_container)
            raw_bl = _val(c_bl) if c_bl is not None else ''
            raw_cost = _val(c_cost)

            # Fila vacía total => saltar
            if not (raw_container or raw_bl or raw_cost):
                continue

            try:
                cost = float(str(raw_cost).replace(',', '.'))
            except Exception:
                raise UserError(_("Costo inválido en fila %s: %s") % (r, raw_cost))

            results.append({'row': r, 'container': raw_container, 'bl': raw_bl, 'cost': cost})

        if not results:
            raise UserError(_("El archivo no contiene filas de datos después del encabezado."))
        return results

    def _find_container_and_import(self, container_name, bl_value):
        """
        Localiza el registro de importación por BL y, si existe, busca el contenedor
        por nombre VINCULADO a esa importación. Retorna (container, import).
        """
        C = self.env[self._container_model].sudo()
        I = self.env[self._import_model].sudo()

        cont = False
        imp = False

        _logger.info("CONTAINER NAME: %s", container_name)
        _logger.info("BL: %s", bl_value)

        # 1) Buscar importación por BL
        if bl_value:
            imports = I.search([(self._container_bl_field, '=', bl_value)])

            # 2) Si hay import, buscar el contenedor por nombre y vinculado a esa importación
            for imp in imports:
                if not cont and container_name and imp:
                    _logger.info("IMPORT ID: %s", getattr(imp, 'id', False))
                    cont = C.with_context(lang='es_419').search([
                        (self._container_name_field, '=', container_name),
                        ('x_studio_many2one_field_15t_1hjr52r0r.id', '=', imp.id)  # vínculo contenedor→importación
                    ], limit=1)
                    if cont:
                        _logger.info("Contenedor encontrado: %s", getattr(cont, 'x_name', False))
                        _logger.info("Importación encontrada: %s", getattr(imp, 'x_name', False))
                        return cont, imp

            return cont, imp

    def _get_clients_from_container(self, container):
        """
        Obtiene clientes únicos desde:
          container.x_studio_order_lines_by_container  --> (o2m intermedio)
            -> line.x_studio_purchase_order_line       --> (m2o purchase.order.line)
            -> purchase.order.line.order_id            --> (m2o purchase.order)
            -> purchase.order.x_studio_client          --> (m2o res.partner)
        """
        InterLine = self.env['x_containerlineorder']
        inter_lines = getattr(container, self._container_lines_o2m, InterLine.browse())
        if not inter_lines:
            return self.env['res.partner']

        po_lines = inter_lines.mapped(self._container_line_po_line_m2o)
        purchase_orders = po_lines.mapped('order_id')
        clients = purchase_orders.mapped('x_studio_client').filtered(lambda p: p and p.id)
        return clients

    def _get_taxes_for_product(self, product):
        company = self.company_id
        taxes = product.taxes_id.filtered(lambda t: t.company_id == company and t.type_tax_use in ('sale', 'none'))
        return [(6, 0, taxes.ids)] if taxes else [(6, 0, [])]

    def _get_or_create_invoice(self, partner, import_rec, container_rec=None):
        Move = self.env['account.move']
        simple_by_partner = (self.grouping_mode == 'by_partner')

        cache = self._context.get('_created_invoices_cache')
        if cache is None:
            cache = {}
            self = self.with_context(_created_invoices_cache=cache)

        key = partner.id if simple_by_partner else (partner.id, import_rec.id if import_rec else False)

        if key in cache:
            inv = cache[key]
            inv.sudo().write({'x_studio_type_import': 'normal'})
            if not simple_by_partner:
                if import_rec:
                    inv.sudo().write({'x_studio_import_id': import_rec.id})
                if container_rec:
                    inv.sudo().write({'x_studio_container_ids': [(4, container_rec.id)]})
            return inv

        vals = {
            'move_type': 'out_invoice',
            'partner_id': partner.id,
            'journal_id': self.journal_id.id,
            'invoice_date': self.invoice_date or date.today(),
            'company_id': self.company_id.id,
            'x_studio_type_import': 'normal',
        }
        if self.force_currency_id:
            vals['currency_id'] = self.force_currency_id.id

        if not simple_by_partner:
            if import_rec:
                vals['x_studio_import_id'] = import_rec.id
            if container_rec:
                vals['x_studio_container_ids'] = [(6, 0, [container_rec.id])]

        inv = Move.with_company(self.company_id).sudo().create(vals)
        cache[key] = inv
        return inv

    def _attach_container_to_move(self, inv, container):
        if self.grouping_mode == 'by_import' and container:
            inv.write({'x_studio_container_ids': [(4, container.id)]})

    def _create_line(self, inv, product, cost, container_rec=None, bl_value=''):
        name_bits = [product.display_name]
        # Siempre detallar contenedor/BL en la descripción de la línea si vienen:
        if container_rec:
            name_bits.append(_("Contenedor: %s") % getattr(container_rec, self._container_name_field, ''))
        if bl_value:
            name_bits.append(_("BL: %s") % bl_value)

        line_vals = {
            'move_id': inv.id,
            'product_id': product.id,
            'name': " - ".join([b for b in name_bits if b]),
            'quantity': 1.0,
            'price_unit': cost,
            'tax_ids': self._get_taxes_for_product(product),
            'account_id': (product.property_account_income_id.id
                           or product.categ_id.property_account_income_categ_id.id
                           or False),
        }
        return self.env['account.move.line'].create(line_vals)

        # =========================================================
        # Helpers período / hash
        # =========================================================

    def _compute_file_hash(self):
        """SHA256 del binario (opcional, para referencia/duplicados)."""
        if not self.file_data:
            return ''
        b = base64.b64decode(self.file_data)
        return hashlib.sha256(b).hexdigest()

    def _extract_period_from_workbook(self, wb):
        """
        Busca 'Periodo: dd/mm/YYYY - dd/mm/YYYY' en cualquier celda
        de la hoja seleccionada (sheet_index). Devuelve (date_from, date_to).
        """
        ws = self._get_worksheet(wb)  # reutiliza tu método existente
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if not cell or not isinstance(cell, str):
                    continue
                m = PERIOD_REGEX.search(cell)
                if m:
                    d1 = _parse_dt_ddmmyyyy(m.group(1))
                    d2 = _parse_dt_ddmmyyyy(m.group(2))
                    if d2 < d1:
                        raise UserError(_("El período del Excel es inválido: la fecha final es anterior a la inicial."))
                    return d1, d2
        raise UserError(_("No se encontró el texto de período en el Excel (ej: 'Periodo: 25/06/2025 - 20/07/2025')."))

    def _extract_period_from_excel(self):
        """Abre el workbook y extrae el período."""
        if openpyxl is None:
            raise UserError(_("Falta la librería 'openpyxl'."))
        try:
            b = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(b), data_only=True)
        except Exception as e:
            raise UserError(_("No se pudo leer el Excel para extraer el período: %s") % e)
        return self._extract_period_from_workbook(wb)

    def _validate_not_duplicate(self, period_start, period_end):
        """
        Valida que NO exista un log con misma tupla (product_id, period_start, period_end, company_id).
        Si existe, lanza UserError.
        """
        Log = self.env['import.container.billing.log']
        exists = Log.search_count([
            ('product_id', '=', self.product_id.id),
            ('period_start', '=', period_start),
            ('period_end', '=', period_end),
            ('company_id', '=', self.company_id.id),
        ])
        if exists:
            raise UserError(_(
                "Este servicio (%(prod)s) ya fue importado para el mismo período (%(d1)s - %(d2)s) "
                "en la compañía seleccionada."
            ) % {
                                'prod': self.product_id.display_name,
                                'd1': period_start,
                                'd2': period_end,
                            })

    # =========================================================
    # Acción principal (extiende tu flujo actual)
    # =========================================================

    def action_process(self):
        """
        Flujo con:
          1) Extraer período del Excel.
          2) Validar duplicado por (product_id, período, company).
          3) Ejecutar proceso de facturación como ya lo tenías.
          4) Registrar log de importación.
        """
        if not self.file_data:
            raise UserError(_("Adjunta el archivo Excel."))
        if not self.product_id:
            raise UserError(_("Selecciona el producto/servicio a facturar."))
        if not self.journal_id:
            raise UserError(_("Selecciona el diario de ventas."))

        # 1) Período
        period_start, period_end = self._extract_period_from_excel()
        self.period_start = period_start
        self.period_end = period_end

        # 2) Duplicado
        self._validate_not_duplicate(period_start, period_end)

        # 3) Proceso original
        self = self.with_context(_created_invoices_cache={})

        rows = self._parse_excel()
        created = []
        errors = []
        Product = self.product_id

        for row in rows:
            try:
                container_rec, import_rec = self._find_container_and_import(row['container'], row['bl'])

                if not container_rec:
                    msg = _("Fila %s: No se encontró contenedor (Nombre: %s, BL: %s).") % (
                        row['row'], row['container'] or '-', row['bl'] or '-'
                    )
                    if self.skip_missing_container:
                        errors.append(msg)
                        continue
                    raise UserError(msg)

                if not import_rec:
                    msg = _("Fila %s: El contenedor '%s' no tiene importación asociada.") % (
                        row['row'], getattr(container_rec, self._container_name_field, '')
                    )
                    if self.skip_missing_import:
                        errors.append(msg)
                        continue
                    raise UserError(msg)

                clients = self._get_clients_from_container(container_rec)
                if not clients:
                    errors.append(_("Fila %s: El contenedor '%s' no tiene clientes asociados vía órdenes de compra.") %
                                  (row['row'], getattr(container_rec, self._container_name_field, '')))
                    continue

                # Prorrateo en partes iguales
                n = len(clients)
                share = round(row['cost'] / n, 2)
                shares = [share] * n
                diff = round(row['cost'] - sum(shares), 2)
                if diff:
                    shares[-1] = round(shares[-1] + diff, 2)

                for idx, partner in enumerate(clients):
                    inv = self._get_or_create_invoice(partner, import_rec, container_rec=container_rec)
                    # reforzar tipo import
                    inv[self._move_type_import_field] = 'normal'
                    # asegurar contenedor
                    self._attach_container_to_move(inv, container_rec)
                    # línea
                    self._create_line(inv, Product, shares[idx], container_rec=container_rec, bl_value=row['bl'])
                    created.append(inv.id)

            except Exception as e:
                _logger.exception("Error procesando fila %s", row.get('row'))
                errors.append(_("Fila %s: %s") % (row.get('row'), e))

        if not self.draft_only and created:
            self.env['account.move'].browse(list(set(created))).action_post()

        # 4) Log (no bloquea si falla)
        try:
            file_hash = self._compute_file_hash()
            self.env['import.container.billing.log'].create({
                'product_id': self.product_id.id,
                'period_start': period_start,
                'period_end': period_end,
                'company_id': self.company_id.id,
                'journal_id': self.journal_id.id,
                'filename': self.file_name or '',
                'file_hash': file_hash,
                'invoices_count': len(set(created)),
                'notes': "\n".join(errors[:50]) if errors else False,
            })
        except Exception as le:
            _logger.warning("No se pudo registrar el log de importación: %s", le, exc_info=True)

        msg = _("Proceso finalizado.\nFacturas afectadas: %s\nErrores: %s") % (len(set(created)), len(errors))
        if errors:
            msg += "\n\n" + "\n".join(errors[:50])
        self.result_message = msg

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {'title': _('Resultado de la importación'), 'message': msg, 'sticky': False}
        }

    # ================== Instrucciones para la vista =================
    instruction_excel = fields.Html(string="Instrucciones", sanitize=False,
                                    default=lambda self: self._default_instruction_excel())
    instruction_client = fields.Html(string="Nota", sanitize=False,
                                     default=lambda self: self._default_instruction_client())
    instruction_linking = fields.Html(string="Vinculación automática", sanitize=False,
                                      default=lambda self: self._default_instruction_linking())

    def _default_instruction_excel(self):
        return """
            <p style="margin:0;">
                Sube un Excel con las columnas para: <b>Contenedor</b>, <b>BL</b> y <b>Costo</b>.<br/>
                Indica debajo los nombres exactos de esas columnas (o la letra de columna).
            </p>
            """

    def _default_instruction_client(self):
        return """
            <p style="margin:0;">
                El cliente se deduce de las líneas del contenedor:
                <code>x_studio_order_lines_by_container → x_studio_purchase_order_line → order_id → x_studio_client</code>.<br/>
                Si hay varios clientes, se crea una factura por cliente y se divide el costo entre ellos.
            </p>
            """

    def _default_instruction_linking(self):
        return """
            <p style="margin:0;">
                Se vincularán <b>x_studio_container_ids</b> y <b>x_studio_import_id</b> en la factura,<br/>
                y se establecerá <b>x_studio_type_import = 'normal'</b>.
            </p>
            """
