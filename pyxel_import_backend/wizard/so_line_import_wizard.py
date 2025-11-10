# -*- coding: utf-8 -*-
import base64
import io
import re
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except Exception:
    openpyxl = None


def _col_to_idx(col_ref):
    if isinstance(col_ref, int):
        return max(col_ref - 1, 0)
    s = (col_ref or '').strip()
    if not s:
        raise UserError(_("Las referencias de columna no pueden estar vacías."))
    if s.isdigit():
        return max(int(s) - 1, 0)
    s = s.upper()
    if not re.fullmatch(r'[A-Z]+', s):
        raise UserError(_("Columna inválida: %s (usa A,B,AA o 1,2,3)") % s)
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _to_float_loose(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    s = re.sub(r'[^\d,.\-]', '', s)
    if s.count(',') and s.count('.'):
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        if s.count(',') and not s.count('.'):
            s = s.replace(',', '.')
    try:
        return float(s)
    except Exception:
        raise UserError(_("No se pudo convertir a número: %s") % v)


def _to_float_safe(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0.0, False
    try:
        return _to_float_loose(v), True
    except Exception:
        return 0.0, False


class SOLineImportWizard(models.TransientModel):
    _name = 'so.line.import.wizard'
    _description = 'Importar líneas a Pedido de Venta desde Excel'

    # === Contexto ===
    sale_id = fields.Many2one(
        'sale.order',
        string="Pedido de Venta",
        required=True,
        domain=[('state', 'in', ('draft', 'sent'))]  # ajusta si necesitas permitir 'sale'
    )

    # === Archivo Excel ===
    file_data = fields.Binary(string="Archivo Excel (.xlsx)", required=True)
    file_name = fields.Char(string="Nombre de archivo")
    sheet_index = fields.Integer(string="Hoja (0=primera)", default=0)

    # === Config tabla de PRODUCTOS ===
    header_row_products = fields.Integer(
        string="Fila encabezado (productos)",
        default=1,
        help="Fila donde está el encabezado. La siguiente fila es la primera de datos."
    )
    max_product_rows = fields.Integer(
        string="Máx. filas productos",
        help="Si se define (>0), corta lectura a esa cantidad de filas."
    )
    empty_product_break = fields.Integer(
        string="Corte por filas vacías",
        default=2,
        help="Si detecta este # de filas seguidas sin producto, asume fin."
    )
    strict_numeric_products = fields.Boolean(
        string="Modo estricto numérico",
        default=False,
        help="Si activo: cantidad no numéricos => error. "
             "Si inactivo: esa fila se omite y se deja nota."
    )

    # Mapeo columnas productos
    col_prod_seq = fields.Char(string="Col. N°", default='A',
                               help="Columna del número de línea (opcional, sólo informativo).")
    col_prod_name = fields.Char(string="Col. Producto", default='B', required=True)
    col_prod_code = fields.Char(string="Col. Código", default='C')
    col_prod_uom = fields.Char(string="Col. U/M", default='D', required=True)
    col_prod_qty = fields.Char(string="Col. Cantidad", default='E', required=True)

    # === Opciones de creación ===
    create_missing_products = fields.Boolean(
        string="Crear productos inexistentes",
        default=True,
        help="Si el producto no existe en Odoo, se creará automáticamente."
    )
    default_product_type = fields.Selection([
        ('product', 'Almacenable'),
        ('consu', 'Consumible'),
        ('service', 'Servicio'),
    ], string="Tipo producto x defecto", default='product')

    default_uom_id = fields.Many2one(
        'uom.uom',
        string="UoM por defecto al crear"
    )

    clear_existing_lines = fields.Boolean(
        string="Reemplazar líneas existentes",
        default=False,
        help="Si activo, borra líneas del pedido antes de importar."
    )

    # Mensaje resultado
    note = fields.Text(readonly=True)

    # -------------------
    # Helpers internos
    # -------------------

    def _get_ws(self, wb):
        idx = self.sheet_index or 0
        sheets = wb.worksheets
        if idx < 0 or idx >= len(sheets):
            raise UserError(_("Índice de hoja fuera de rango. El archivo tiene %s hoja(s).") % len(sheets))
        return sheets[idx]

    def _find_uom(self, uom_name):
        if not uom_name:
            return self.default_uom_id
        UoM = self.env['uom.uom'].sudo()
        u = UoM.search([('name', '=ilike', uom_name)], limit=1)
        if not u:
            u = UoM.search([('name', 'ilike', uom_name)], limit=1)
        return u or self.default_uom_id

    def _get_or_create_product(self, code, name, uom):
        Product = self.env['product.product'].sudo()

        # 1) buscar primero por código
        prod = False
        if code:
            prod = Product.search([('default_code', '=', code)], limit=1)

        # 2) si no, por nombre EXACTO
        if not prod and name:
            prod = Product.search([('name', '=', name)], limit=1)

        # ========== caso: el producto YA existe ==========
        if prod:
            # si el Excel trae una U/M, validamos contra la del producto
            if uom:
                # en compras Odoo mira sobre todo uom_po_id
                prod_uom = prod.uom_po_id or prod.uom_id
                if prod_uom and prod_uom.category_id != uom.category_id:
                    # categorías distintas -> no son compatibles
                    raise UserError(_(
                        "El producto '%(prod)s' ya existe con la U/M '%(p_uom)s' "
                        "y en el Excel viene la U/M '%(x_uom)s'. "
                        "Debes usar la misma U/M o corregir el producto."
                    ) % {
                                        'prod': prod.display_name,
                                        'p_uom': prod_uom.display_name,
                                        'x_uom': uom.display_name,
                                    })
                # mismas categorías pero distinto id → también lo puedes forzar a que se use la del producto
                # o lanzar error; aquí lanzo error igual para que no haya sorpresas:
                if prod_uom and prod_uom.id != uom.id:
                    raise UserError(_(
                        "El producto '%(prod)s' usa la U/M '%(p_uom)s', "
                        "pero el Excel trae '%(x_uom)s'. Usa la U/M del producto."
                    ) % {
                                        'prod': prod.display_name,
                                        'p_uom': prod_uom.display_name,
                                        'x_uom': uom.display_name,
                                    })
            return prod

        # ========== caso: NO existe y vamos a crearlo ==========
        if not self.create_missing_products:
            raise UserError(_("El producto '%s' no existe y la creación automática está desactivada.") % name)

        if not uom:
            raise UserError(_(
                "No se pudo determinar la Unidad de Medida para el producto '%s'. "
                "Configura una U/M por defecto en el wizard o en el Excel."
            ) % name)

        tmpl_vals = {
            'name': name,
            'default_code': code,
            'type': self.default_product_type,
            'uom_id': uom.id,
            'uom_po_id': uom.id,
            'taxes_id': [(6, 0, [])],
            'supplier_taxes_id': [(6, 0, [])],
        }
        tmpl = self.env['product.template'].sudo().create(tmpl_vals)
        return tmpl.product_variant_id

    def _parse_product_lines(self, ws):
        start = max(self.header_row_products or 1, 1) + 1
        end = ws.max_row

        if self.max_product_rows and self.max_product_rows > 0:
            end = min(end, start + self.max_product_rows - 1)

        c_name = _col_to_idx(self.col_prod_name)
        c_code = _col_to_idx(self.col_prod_code) if self.col_prod_code else None
        c_uom = _col_to_idx(self.col_prod_uom)
        c_qty = _col_to_idx(self.col_prod_qty)
        # c_price = _col_to_idx(self.col_prod_price)

        lines = []
        notes = []
        empty_streak = 0

        for row_idx in range(start, end + 1):

            def _cell(ci):
                return ws.cell(row=row_idx, column=ci + 1).value if ci is not None else None

            raw_name = _cell(c_name)
            name = (raw_name or '').strip() if isinstance(raw_name, str) else raw_name

            if not name:
                empty_streak += 1
                if self.empty_product_break and empty_streak >= self.empty_product_break:
                    break
                continue
            empty_streak = 0

            raw_uom = _cell(c_uom)
            uom_name = (raw_uom or '').strip() if isinstance(raw_uom, str) else raw_uom

            qty_val, qty_ok = _to_float_safe(_cell(c_qty))

            if self.strict_numeric_products and (not qty_ok ):
                raise UserError(
                    _("Fila %s: Cantidad/Precio no son numéricos (%s / %s).") %
                    (row_idx, _cell(c_qty))
                )

            if not qty_ok:
                notes.append(
                    _("Fila %s omitida por datos no numéricos. Cantidad=%s") %
                    (row_idx, _cell(c_qty))
                )
                continue

            code_val = _cell(c_code) if c_code is not None else ''

            lines.append({
                'row': row_idx,
                'name': name,
                'code': code_val or '',
                'uom_name': uom_name or '',
                'qty': qty_val,
                'price': 0,
            })

        return lines, notes

    def _create_so_line(self, SO, prod, uom, qty, price, name):
        self.env['sale.order.line'].sudo().create({
            'order_id': SO.id,
            'product_id': prod.id,
            'name': name,
            'product_uom_qty': qty,
            'price_unit': price,
            'product_uom': uom.id if uom else prod.uom_id.id,
            'tax_id': [(6, 0, [])],
        })

    # -------------------
    # Acción principal
    # -------------------

    def action_import(self):
        if not self.sale_id:
            raise UserError(_("Debes indicar el Pedido de Venta."))
        if not self.file_data:
            raise UserError(_("Adjunta el archivo Excel."))
        if openpyxl is None:
            raise UserError(_("Falta la librería 'openpyxl' en el servidor."))

        try:
            data = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("No se pudo leer el Excel: %s") % e)

        ws = self._get_ws(wb)
        SO = self.sale_id.sudo()

        if self.clear_existing_lines and SO.order_line:
            SO.order_line.unlink()

        product_lines, notes_prod = self._parse_product_lines(ws)

        created_count = 0
        for pl in product_lines:
            uom = self._find_uom(pl['uom_name'])
            prod = self._get_or_create_product(pl['code'], pl['name'], uom)
            self._create_so_line(
                SO,
                prod,
                uom,
                qty=pl['qty'],
                price=pl['price'],
                name=pl['name']
            )
            created_count += 1

        msg = _("Se importaron %(cnt)s líneas de productos en el Pedido de Venta %(so)s.") % {
            'cnt': created_count,
            'so': SO.name,
        }
        if notes_prod:
            msg += "\n" + "\n".join(notes_prod)

        self.note = msg

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación completada'),
                'message': msg,
                'sticky': False,
            }
        }
