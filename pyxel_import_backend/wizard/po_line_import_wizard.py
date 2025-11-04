# -*- coding: utf-8 -*-
import base64
import io
import re
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except Exception:
    openpyxl = None


# -----------------------
# Utilidades comunes
# -----------------------

def _col_to_idx(col_ref):
    """Convierte 'A'/'B'/'AA' o '1'/'2' -> índice 0-based (int)."""
    if isinstance(col_ref, int):
        return max(col_ref - 1, 0)

    s = (col_ref or '').strip()
    if not s:
        raise UserError(_("Las referencias de columna no pueden estar vacías."))

    if s.isdigit():
        return max(int(s) - 1, 0)

    s = s.upper()
    if not re.fullmatch(r'[A-Z]+', s):
        raise UserError(_("Columna inválida: %s (usa A,B,AA o 1,2,3)") % s)

    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _normalize_txt(txt):
    """Normaliza texto básico para comparar con etiquetas ('FLETE', 'SEGURO', etc.)."""
    if not txt:
        return ''
    t = str(txt).upper()
    t = (
        t.replace('Á', 'A').replace('É', 'E').replace('Í', 'I')
         .replace('Ó', 'O').replace('Ú', 'U').replace('Ñ', 'N')
    )
    return t.strip()


def _to_float_loose(v):
    """
    Convierte algo como '1,234.56', '1.234,56', '$5.00' -> float
    """
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    s = re.sub(r'[^\d,.\-]', '', s)

    if s.count(',') and s.count('.'):
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        if s.count(',') and not s.count('.'):
            s = s.replace(',', '.')

    try:
        return float(s)
    except Exception:
        raise UserError(_("No se pudo convertir a número: %s") % v)


def _to_float_safe(v):
    """Devuelve (float_value, es_valido_bool)."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0.0, False
    try:
        return _to_float_loose(v), True
    except Exception:
        return 0.0, False


# -----------------------
# Wizard PO
# -----------------------

class POLineImportWizard(models.TransientModel):
    _name = 'po.line.import.wizard'
    _description = 'Importar líneas a Orden de Compra desde Excel'

    # === Contexto ===
    purchase_id = fields.Many2one(
        'purchase.order',
        string="Orden de Compra",
        required=True,
        domain=[('state', 'in', ('draft', 'sent'))]
    )

    # === Archivo Excel ===
    file_data = fields.Binary(string="Archivo Excel (.xlsx)", required=True)
    file_name = fields.Char(string="Nombre de archivo")
    sheet_index = fields.Integer(string="Hoja (0=primera)", default=0)

    # === Config tabla de PRODUCTOS ===
    header_row_products = fields.Integer(
        string="Fila encabezado (productos)",
        default=1,
        help="Fila donde está el encabezado de la tabla de productos. La siguiente fila es la primera de datos."
    )
    max_product_rows = fields.Integer(
        string="Máx. filas productos",
        help="Si se define (>0), corta a esta cantidad de filas de productos desde el encabezado+1."
    )
    empty_product_break = fields.Integer(
        string="Corte por filas vacías",
        default=2,
        help="Si detecta este número de filas consecutivas sin producto, asume fin de la tabla de productos."
    )
    strict_numeric_products = fields.Boolean(
        string="Modo estricto numérico (productos)",
        default=False,
        help="Si está activo y Cantidad/Precio no son numéricos, lanza error. "
             "Si está desactivado, esa fila se ignora y se deja nota."
    )

    # Mapeo columnas productos
    col_prod_seq = fields.Char(string="Col. N°", default='A')
    col_prod_name = fields.Char(string="Col. Producto", default='B', required=True)
    col_prod_code = fields.Char(string="Col. Código", default='C')
    col_prod_uom = fields.Char(string="Col. U/M", default='D', required=True)
    col_prod_qty = fields.Char(string="Col. Cantidad", default='E', required=True)
    col_prod_price = fields.Char(string="Col. Precio Unitario", default='F', required=True)

    # === Config tabla de SERVICIOS ===
    header_row_services = fields.Integer(
        string="Fila inicial servicios",
        default=2,
        help="Fila donde empiezan los servicios en la tabla lateral (sin incluir encabezado 'Servicio | Importe')."
    )
    max_service_rows = fields.Integer(
        string="Máx. filas servicios",
        default=10,
        help="Límite de filas a leer en la tabla lateral de servicios."
    )
    empty_service_break = fields.Integer(
        string="Corte filas vacías (servicios)",
        default=2,
        help="Si detectamos este # de filas seguidas sin servicio, dejamos de leer."
    )
    strict_numeric_services = fields.Boolean(
        string="Modo estricto numérico (servicios)",
        default=False,
        help="Si está activo y el Importe del servicio no es numérico, error. "
             "Si está desactivado, se omite esa fila."
    )
    col_srv_name = fields.Char(string="Col. Servicio", default='J')
    col_srv_amount = fields.Char(string="Col. Importe", default='K')

    # === Opciones de creación ===
    create_missing_products = fields.Boolean(
        string="Crear productos inexistentes",
        default=True,
    )
    default_product_type = fields.Selection([
        ('product', 'Almacenable'),
        ('consu', 'Consumible'),
        ('service', 'Servicio'),
    ], string="Tipo producto x defecto", default='product')

    default_uom_id = fields.Many2one(
        'uom.uom',
        string="UoM por defecto al crear",
    )

    clear_existing_lines = fields.Boolean(
        string="Reemplazar líneas existentes",
        default=False,
        help="Si está activo, elimina todas las líneas actuales de la OC antes de importar."
    )

    # Productos que representan cada servicio
    product_fob_id = fields.Many2one(
        'product.product',
        string="Servicio: Gasto FOB",
        domain=[('detailed_type', '=', 'service')]
    )
    product_freight_id = fields.Many2one(
        'product.product',
        string="Servicio: Flete",
        domain=[('detailed_type', '=', 'service')]
    )
    product_insurance_id = fields.Many2one(
        'product.product',
        string="Servicio: Seguro",
        domain=[('detailed_type', '=', 'service')]
    )
    product_other_id = fields.Many2one(
        'product.product',
        string="Servicio: Otros gastos",
        domain=[('detailed_type', '=', 'service')]
    )

    note = fields.Text(readonly=True)

    # -------------------
    # Helpers internos
    # -------------------

    def _get_ws(self, wb):
        idx = self.sheet_index or 0
        sheets = wb.worksheets
        if idx < 0 or idx >= len(sheets):
            raise UserError(_("Índice de hoja fuera de rango. El archivo tiene %s hoja(s).") % len(sheets))
        return sheets[idx]

    def _find_uom(self, uom_name):
        if not uom_name:
            return self.default_uom_id
        UoM = self.env['uom.uom'].sudo()
        u = UoM.search([('name', '=ilike', uom_name)], limit=1)
        if not u:
            u = UoM.search([('name', 'ilike', uom_name)], limit=1)
        return u or self.default_uom_id

    def _get_or_create_product(self, name, uom):
        Product = self.env['product.product'].sudo()
        prod = Product.search([('name', '=ilike', name)], limit=1)
        if not prod:
            prod = Product.search([('name', 'ilike', name)], limit=1)

        if prod:
            return prod

        if not self.create_missing_products:
            raise UserError(_("El producto '%s' no existe y la creación automática está desactivada.") % name)

        if not uom:
            raise UserError(_("No se puede crear el producto '%s' porque no se obtuvo una U/M válida.") % name)

        tmpl_vals = {
            'name': name,
            'type': self.default_product_type,
            'uom_id': uom.id,
            'uom_po_id': uom.id,
            'taxes_id': [(6, 0, [])],
            'supplier_taxes_id': [(6, 0, [])],
        }
        tmpl = self.env['product.template'].sudo().create(tmpl_vals)
        return tmpl.product_variant_id

    # -------------------
    # Parseo de productos
    # -------------------

    def _parse_product_lines(self, ws):
        """Lee el bloque principal (Producto, Código, U/M, Cantidad, Precio)."""
        start = max(self.header_row_products or 1, 1) + 1
        end = ws.max_row

        if self.max_product_rows and self.max_product_rows > 0:
            end = min(end, start + self.max_product_rows - 1)

        c_name = _col_to_idx(self.col_prod_name)
        c_code = _col_to_idx(self.col_prod_code) if self.col_prod_code else None
        c_uom = _col_to_idx(self.col_prod_uom)
        c_qty = _col_to_idx(self.col_prod_qty)
        c_price = _col_to_idx(self.col_prod_price)

        lines = []
        notes = []
        empty_streak = 0

        for row_idx in range(start, end + 1):

            def _cell(ci):
                return ws.cell(row=row_idx, column=ci + 1).value if ci is not None else None

            raw_name = _cell(c_name)
            name = (raw_name or '').strip() if isinstance(raw_name, str) else raw_name

            # si no hay producto, contamos vacíos y quizás cortamos
            if not name:
                empty_streak += 1
                if self.empty_product_break and empty_streak >= self.empty_product_break:
                    break
                continue
            empty_streak = 0

            # UoM: forzar a string siempre
            raw_uom = _cell(c_uom)
            uom_name = str(raw_uom or '').strip()

            qty_val, qty_ok = _to_float_safe(_cell(c_qty))
            price_val, price_ok = _to_float_safe(_cell(c_price))

            if self.strict_numeric_products and (not qty_ok or not price_ok):
                raise UserError(
                    _("Fila %s: Cantidad/Precio no son numéricos (%s / %s).") %
                    (row_idx, _cell(c_qty), _cell(c_price))
                )

            if not qty_ok or not price_ok:
                notes.append(
                    _("Fila %s omitida por datos no numéricos. Cantidad=%s Precio=%s") %
                    (row_idx, _cell(c_qty), _cell(c_price))
                )
                continue

            # aquí el fix: lo que venga de la col de código lo paso SIEMPRE a str
            raw_code = _cell(c_code) if c_code is not None else ''
            code_val = str(raw_code or '').strip()

            lines.append({
                'row': row_idx,
                'name': name,
                'code': code_val,
                'uom_name': uom_name,
                'qty': qty_val,
                'price': price_val,
            })

        return lines, notes

    # -------------------
    # Parseo de servicios
    # -------------------

    def _parse_service_lines(self, ws):
        start = max(self.header_row_services or 1, 1)
        end = ws.max_row

        if self.max_service_rows and self.max_service_rows > 0:
            end = min(end, start + self.max_service_rows - 1)

        c_srv_name = _col_to_idx(self.col_srv_name)
        c_srv_amount = _col_to_idx(self.col_srv_amount)

        lines = []
        notes = []
        empty_streak = 0

        for row_idx in range(start, end + 1):

            def _cell(ci):
                return ws.cell(row=row_idx, column=ci + 1).value if ci is not None else None

            raw_srv = _cell(c_srv_name)
            srv_name = (raw_srv or '').strip() if isinstance(raw_srv, str) else raw_srv

            raw_amt = _cell(c_srv_amount)
            amt_val, amt_ok = _to_float_safe(raw_amt)

            if not srv_name and (raw_amt is None or str(raw_amt).strip() == ''):
                empty_streak += 1
                if self.empty_service_break and empty_streak >= self.empty_service_break:
                    break
                continue
            empty_streak = 0

            if self.strict_numeric_services and not amt_ok:
                raise UserError(
                    _("Fila servicio %s: importe no numérico (%s).") %
                    (row_idx, raw_amt)
                )

            if not srv_name:
                notes.append(_("Fila servicio %s omitida: sin nombre de servicio.") % row_idx)
                continue

            if not amt_ok:
                notes.append(_("Fila servicio %s omitida: importe no numérico (%s).") %
                             (row_idx, raw_amt))
                continue

            lines.append({
                'row': row_idx,
                'service': srv_name,
                'amount': amt_val,
            })

        return lines, notes

    # -------------------
    # Crear / actualizar líneas en PO
    # -------------------

    def _create_po_product_line(self, PO, prod, uom, qty, price, name):
        self.env['purchase.order.line'].sudo().create({
            'order_id': PO.id,
            'product_id': prod.id,
            'name': name,
            'product_qty': qty,
            'price_unit': price,
            'product_uom': (uom.id if uom else prod.uom_po_id.id),
        })

    def _create_po_service_line(self, PO, srv_name, amount):
        norm = _normalize_txt(srv_name)

        if 'FOB' in norm and self.product_fob_id:
            service_prod = self.product_fob_id
        elif 'FLETE' in norm and self.product_freight_id:
            service_prod = self.product_freight_id
        elif 'SEGURO' in norm and self.product_insurance_id:
            service_prod = self.product_insurance_id
        elif self.product_other_id:
            service_prod = self.product_other_id
        else:
            return _("No se creó línea de servicio '%s' (importe %s) porque no hay producto asignado.") % (
                srv_name, amount
            )

        self.env['purchase.order.line'].sudo().create({
            'order_id': PO.id,
            'product_id': service_prod.id,
            'name': service_prod.display_name,
            'product_qty': 1.0,
            'price_unit': amount,
            'product_uom': service_prod.uom_po_id.id or service_prod.uom_id.id,
        })
        return False

    # -------------------
    # Acción principal
    # -------------------

    def action_import(self):
        if not self.purchase_id:
            raise UserError(_("Debes indicar la Orden de Compra."))
        if not self.file_data:
            raise UserError(_("Adjunta el archivo Excel."))
        if openpyxl is None:
            raise UserError(_("Falta la librería 'openpyxl' en el servidor."))

        try:
            data = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(_("No se pudo leer el Excel: %s") % e)

        ws = self._get_ws(wb)
        PO = self.purchase_id.sudo()

        # 1) parseo
        product_lines, notes_prod = self._parse_product_lines(ws)
        service_lines, notes_srv = self._parse_service_lines(ws)

        # si NO se marca "reemplazar", vamos a ACTUALIZAR
        not_matched = []
        uom_mismatch = []
        created_count = 0
        updated_count = 0

        if self.clear_existing_lines and PO.order_line:
            PO.order_line.unlink()

        # índice de líneas existentes (solo cuando NO limpiamos)
        existing_by_name = {}
        existing_by_code = {}
        if not self.clear_existing_lines:
            for line in PO.order_line:
                # por nombre
                key_name = (line.product_id and line.product_id.name) or line.name or ''
                if key_name:
                    existing_by_name[key_name.strip().upper()] = line
                # por código
                if line.product_id and line.product_id.default_code:
                    existing_by_code[line.product_id.default_code.strip().upper()] = line

        # 2) procesar productos
        for pl in product_lines:
            excel_name = pl['name']
            excel_code = (pl['code'] or '').strip().upper()
            uom = self._find_uom(pl['uom_name'])

            if not self.clear_existing_lines:
                # modo actualizar
                line_found = False
                target_line = None

                if excel_code:
                    target_line = existing_by_code.get(excel_code)

                if not target_line:
                    target_line = existing_by_name.get(excel_name.strip().upper())

                if target_line:
                    # validar UoM
                    if pl['uom_name'] and target_line.product_uom and \
                            target_line.product_uom.name.strip().upper() != pl['uom_name'].strip().upper():
                        uom_mismatch.append(
                            _("Línea '%(name)s': U/M del Excel (%(u1)s) difiere de la OC (%(u2)s).") % {
                                'name': excel_name,
                                'u1': pl['uom_name'],
                                'u2': target_line.product_uom.name,
                            }
                        )
                    target_line.write({
                        'price_unit': pl['price'],
                        'product_qty': pl['qty'],
                    })
                    updated_count += 1
                    line_found = True

                if not line_found:
                    not_matched.append(excel_name)
                continue

            # modo crear (clear_existing_lines = True)
            prod = self._get_or_create_product(excel_name, uom)
            self._create_po_product_line(
                PO, prod, uom,
                qty=pl['qty'],
                price=pl['price'],
                name=excel_name,
            )
            created_count += 1

        # 3) procesar servicios (se crean siempre)
        for sl in service_lines:
            warn = self._create_po_service_line(PO, sl['service'], sl['amount'])
            if warn:
                notes_srv.append(warn)

        # 4) mensaje final
        msg_parts = []
        if self.clear_existing_lines:
            msg_parts.append(_("Se importaron %(n)s líneas de productos en la OC %(po)s.") % {
                'n': created_count,
                'po': PO.name,
            })
        else:
            msg_parts.append(_("Se actualizaron %(n)s líneas existentes en la OC %(po)s.") % {
                'n': updated_count,
                'po': PO.name,
            })

        if not self.clear_existing_lines and not_matched:
            msg_parts.append(_("No se encontraron en la orden estos productos del Excel:"))
            for nm in not_matched:
                msg_parts.append(" - %s" % nm)

        if uom_mismatch:
            msg_parts.append(_("Diferencias de unidad de medida detectadas:"))
            msg_parts.extend(uom_mismatch)

        if notes_prod:
            msg_parts.append("\n".join(notes_prod))
        if notes_srv:
            msg_parts.append("\n".join(notes_srv))

        msg = "\n".join(msg_parts)
        self.note = msg

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación completada'),
                'message': msg,
                'sticky': False
            }
        }
